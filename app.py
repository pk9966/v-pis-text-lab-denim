""import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from difflib import SequenceMatcher

st.set_page_config(page_title="Výpis dle kritérií")
st.title("Vyhledání dle klíče (PM+LM OP1)")

col1, col2 = st.columns(2)
with col1:
    lab_file = st.file_uploader("Nahraj laboratorní deník (XLSX, list 'Evidence zkoušek zhotovitele')", type="xlsx", key="lab")
with col2:
    klic_file = st.file_uploader("Nahraj klíč (XLSX se seznamem zkoušek)", type="xlsx", key="klic")

cislo_objektu_input = st.text_input("Volitelné: Číslo objektu pro filtrování (např. 209)").strip()
ladi = st.checkbox("🔧 Zobrazit ladicí informace")

if lab_file and klic_file:
    lab_bytes = lab_file.read()
    klic_bytes = klic_file.read()

    df = pd.read_excel(io.BytesIO(lab_bytes), sheet_name="Evidence zkoušek zhotovitele")
    workbook = load_workbook(io.BytesIO(klic_bytes))

    try:
        klic_df = pd.read_excel(io.BytesIO(klic_bytes), sheet_name="seznam zkoušek PM+LM OP1", header=None)
    except Exception as e:
        st.error(f"Chyba při čtení klíče: {e}")
        st.stop()

    df.columns.values[10] = "K"  # konstrukční prvek
    df.columns.values[13] = "N"  # druh zkoušky
    df.columns.values[7] = "H"   # staničení
    df.columns.values[2] = "C"   # číslo objektu

    st.markdown("""
    ### 🔍 Pravidla pro vyhledávání
    - **Pravidlo 1**: Konstrukční prvek (sloupec K) obsahuje zadaný text (např. „zásyp“ → „zásyp základů za opěrou“)
    - **Pravidlo 2**: Druh zkoušky (sloupec N) obsahuje alespoň jednu hodnotu ze seznamu zadaného v klíči
    - **Pravidlo 3**: Číslo objektu (sloupec C) – pokud je zadáno, musí být obsaženo jako podřetězec (např. 209 v "SO 209")
    - **Pravidlo 4**: Staničení (sloupec H) – pokud je uvedeno v klíči, alespoň jedna hodnota musí být obsažena
    """)

    def contains_relaxed(text, keyword):
        return all(k in text for k in keyword.split())

    total_matches = 0
    all_matched_rows = []

    for row_idx in range(2, 11):  # řádky 2 až 10
        if pd.isna(klic_df.iloc[row_idx - 1, 1]):
            continue

        konstrukce = str(klic_df.iloc[row_idx - 1, 1]).strip().lower().replace("-", " ")
        zkouska = str(klic_df.iloc[row_idx - 1, 2]).strip().lower().replace("-", " ")
        stanice = str(klic_df.iloc[row_idx - 1, 3]).strip().lower()

        zkousky = [z.strip().lower() for z in zkouska.split(",") if z.strip()]
        stanice_list = [s.strip().lower() for s in stanice.split(",") if s.strip()]

        local_match_count = 0

        for _, row in df.iterrows():
            text_konstrukce = str(row.get("K", "")).lower().replace("-", " ").strip()
            text_zkouska = str(row.get("N", "")).lower().replace("-", " ").strip()
            text_stanice = str(row.get("H", "")).lower().strip()
            text_objekt = str(row.get("C", "")).replace(" ", "").lower()
            input_objekt = cislo_objektu_input.replace(" ", "").lower()

            konstrukce_ok = contains_relaxed(text_konstrukce, konstrukce)
            zkouska_ok = any(z in text_zkouska for z in zkousky)
            stanice_ok = any(s in text_stanice for s in stanice_list) if stanice_list else True
            objekt_ok = True
            if cislo_objektu_input:
                objekt_ok = input_objekt in text_objekt

            if konstrukce_ok and zkouska_ok and stanice_ok and objekt_ok:
                local_match_count += 1
                all_matched_rows.append({
                    "D (Datum odběru)": row.iloc[5],
                    "E (Staničení)": row["H"],
                    "H (Konstrukční část)": row.iloc[9],
                    "J (Konstrukční prvek)": row["K"],
                    "K (Materiál)": row.iloc[11],
                    "L (Datum zkoušky)": row.iloc[12],
                    "N (Druh zkoušky)": row["N"],
                    "O (Požadovaná hodnota)": row.iloc[14],
                    "P (Naměřená hodnota)": row.iloc[15],
                    "Q (Hodnocení)": row.iloc[16],
                })
            elif ladi:
                fail_reasons = []
                if not konstrukce_ok:
                    fail_reasons.append("konstrukce")
                if not zkouska_ok:
                    fail_reasons.append("zkouška")
                if not stanice_ok:
                    fail_reasons.append("staničení")
                if not objekt_ok:
                    fail_reasons.append("číslo objektu")

                st.write(f"🚫 Řádek: konstrukce='{text_konstrukce}', zkouška='{text_zkouska}', staničení='{text_stanice}', objekt='{text_objekt}'")
                st.write("❌ Nesplněno:", ", ".join(fail_reasons))

        try:
            ws = workbook["PM - OP1"]
            ws[f"D{row_idx}"] = local_match_count
            pozadovano = klic_df.iloc[row_idx - 1, 4]  # sloupec E v klíči
            if pd.notna(pozadovano):
                ws[f"E{row_idx}"] = "Vyhovující" if local_match_count >= int(pozadovano) else f"Chybí {abs(int(pozadovano) - local_match_count)} zk."
            total_matches += local_match_count
        except:
            st.warning(f"List 'PM - OP1' neobsahuje řádek {row_idx}")

    if all_matched_rows:
        st.subheader("🔎 Nalezené odpovídající řádky")
        st.dataframe(pd.DataFrame(all_matched_rows))

    output = io.BytesIO()
    workbook.save(output)

    st.success(f"✅ Celkem nalezeno {total_matches} shod. Výsledky zapsány do listu 'PM - OP1'")
    st.download_button(
        label="📥 Stáhnout aktualizovaný soubor",
        data=output.getvalue(),
        file_name="klic_vyhodnoceny.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
